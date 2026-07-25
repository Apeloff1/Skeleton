"""
gameforge.jeeves.artifacts — Jeeves multi-format creation engine.

Deterministic, FREE (local, no LLM cost) generators that let Jeeves reply in
many forms in a SINGLE parse: PDF, spreadsheet (XLSX), chart variations
(bar/line/pie/scatter), a node graph, and rendered visuals (infographic).

Every generator returns a base64 payload + mime type so the API can hand a
bundle straight to the client. matplotlib runs on the headless Agg backend.
"""
from __future__ import annotations

import base64
import io
from typing import Any, Dict, List, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import networkx as nx  # noqa: E402

# GameForge palette — NO cyan/teal (banned).
_PALETTE = ["#7c3aed", "#22c55e", "#f59e0b", "#ef4444", "#3b82f6", "#ec4899", "#a855f7", "#10b981"]
_BG = "#0b1220"
_FG = "#e2e8f0"


def _fig_to_b64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight", facecolor=_BG)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def _style(ax):
    ax.set_facecolor(_BG)
    for spine in ax.spines.values():
        spine.set_color("#334155")
    ax.tick_params(colors=_FG, labelsize=8)
    ax.title.set_color(_FG)
    ax.xaxis.label.set_color(_FG)
    ax.yaxis.label.set_color(_FG)


# ── charts ─────────────────────────────────────────────────────
def make_chart(kind: str, labels: List[str], values: List[float], title: str = "") -> Dict:
    fig, ax = plt.subplots(figsize=(5, 3))
    _style(ax)
    kind = (kind or "bar").lower()
    n = len(values)
    colors = [_PALETTE[i % len(_PALETTE)] for i in range(max(n, 1))]
    if kind == "line":
        ax.plot(labels, values, color=_PALETTE[0], marker="o", linewidth=2)
    elif kind == "pie":
        ax.pie(values, labels=labels, colors=colors, autopct="%1.0f%%",
               textprops={"color": _FG, "fontsize": 8})
    elif kind == "scatter":
        ax.scatter(range(n), values, c=colors, s=60)
        ax.set_xticks(range(n)); ax.set_xticklabels(labels)
    else:  # bar
        ax.bar(labels, values, color=colors)
    ax.set_title(title or kind.title())
    fig.patch.set_facecolor(_BG)
    return {"type": "chart", "kind": kind, "title": title,
            "mime": "image/png", "base64": _fig_to_b64(fig)}


def make_chart_variations(labels: List[str], values: List[float], title: str = "") -> List[Dict]:
    """Produce several chart variations of the SAME data in one parse."""
    return [make_chart(k, labels, values, f"{title} · {k}".strip(" ·"))
            for k in ("bar", "line", "pie", "scatter")]


# ── graph (nodes/edges) ────────────────────────────────────────
def make_graph(nodes: List[str], edges: List[List[str]], title: str = "Knowledge Graph") -> Dict:
    G = nx.DiGraph()
    G.add_nodes_from(nodes)
    for e in edges:
        if len(e) >= 2:
            G.add_edge(e[0], e[1])
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.set_facecolor(_BG); fig.patch.set_facecolor(_BG)
    pos = nx.spring_layout(G, seed=7) if G.number_of_nodes() else {}
    nx.draw_networkx_nodes(G, pos, ax=ax, node_color=_PALETTE[0], node_size=900, alpha=0.9)
    nx.draw_networkx_edges(G, pos, ax=ax, edge_color="#64748b", arrows=True, arrowsize=14)
    nx.draw_networkx_labels(G, pos, ax=ax, font_color="#ffffff", font_size=8)
    ax.set_title(title, color=_FG); ax.axis("off")
    return {"type": "graph", "title": title, "nodes": len(nodes), "edges": G.number_of_edges(),
            "mime": "image/png", "base64": _fig_to_b64(fig)}


# ── visual (infographic panel) ─────────────────────────────────
def make_visual(title: str, bullets: List[str], metric: Optional[Dict] = None) -> Dict:
    fig, ax = plt.subplots(figsize=(5, 3.2))
    ax.set_facecolor(_BG); fig.patch.set_facecolor(_BG); ax.axis("off")
    ax.text(0.04, 0.92, title, color=_PALETTE[0], fontsize=15, fontweight="bold", va="top")
    y = 0.72
    for b in bullets[:6]:
        ax.text(0.06, y, f"▸ {b}", color=_FG, fontsize=10, va="top")
        y -= 0.12
    if metric:
        ax.text(0.96, 0.92, str(metric.get("value", "")), color=_PALETTE[1],
                fontsize=26, fontweight="bold", ha="right", va="top")
        ax.text(0.96, 0.74, str(metric.get("label", "")), color="#94a3b8",
                fontsize=9, ha="right", va="top")
    return {"type": "visual", "title": title, "mime": "image/png", "base64": _fig_to_b64(fig)}


# ── spreadsheet (XLSX) ─────────────────────────────────────────
def make_spreadsheet(title: str, headers: List[str], rows: List[List[Any]]) -> Dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet")[:28]
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="7C3AED")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = fill
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO(); wb.save(buf)
    return {"type": "spreadsheet", "title": title,
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"{(title or 'sheet').replace(' ', '_')[:40]}.xlsx",
            "base64": base64.b64encode(buf.getvalue()).decode()}


# ── PDF ────────────────────────────────────────────────────────
def make_pdf(title: str, sections: List[Dict], image_b64: Optional[str] = None) -> Dict:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LETTER, title=title)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("H", parent=styles["Title"], textColor=colors.HexColor("#7c3aed"))
    sub = ParagraphStyle("S", parent=styles["Heading2"], textColor=colors.HexColor("#22c55e"))
    body = styles["BodyText"]
    story = [Paragraph(title, h), Spacer(1, 0.2 * inch)]
    for sec in sections:
        if sec.get("heading"):
            story.append(Paragraph(sec["heading"], sub))
        if sec.get("body"):
            story.append(Paragraph(sec["body"], body))
        story.append(Spacer(1, 0.12 * inch))
    if image_b64:
        try:
            story.append(RLImage(io.BytesIO(base64.b64decode(image_b64)), width=4.5 * inch, height=2.7 * inch))
        except Exception:  # noqa: BLE001
            pass
    doc.build(story)
    return {"type": "pdf", "title": title, "mime": "application/pdf",
            "filename": f"{(title or 'doc').replace(' ', '_')[:40]}.pdf",
            "base64": base64.b64encode(buf.getvalue()).decode()}


__all__ = ["make_chart", "make_chart_variations", "make_graph", "make_visual",
           "make_spreadsheet", "make_pdf"]
